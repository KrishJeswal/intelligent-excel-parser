from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook


OUT_DIR = Path(__file__).resolve().parents[1] / "sample_files"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def make_clean_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean"
    ws.append(["Date", "Coal Consumption (MT)", "Coal GCV (kcal/kg)", "Steam Generation (T/hr)", "Power Generation (MWh)", "Water Consumption (KL)"])
    ws.append(["2026-02-20", 1200.0, 4500, 110.5, 85.2, 320])
    ws.append(["2026-02-21", 1250.0, 4550, 112.0, 86.1, 310])
    wb.save(OUT_DIR / "clean_data.xlsx")


def make_messy_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Messy"
    ws.append(["Daily Energy Report - Unit 1"])  
    ws.append([None])  
    ws.append(["DATE", "Coal Consump (AFBC Boiler 1)", "GCV kcal/kg", "Steam Gen T/hr (AFBC-1)", "Pwr Gen (TG1)", "Eff % AFBC-1", "Comments"])
    ws.append(["20-02-2026", "1,234.56", "4,560", "110.5", "85.2", "45%", "ok"])
    ws.append(["21-02-2026", "1,250", "4550", "112", "86.1", "YES", "N/A"])
    ws.append(["22-02-2026", "-", "4,500", None, "86.8", "101%", "check eff"])  
    wb.save(OUT_DIR / "messy_data.xlsx")


def make_multi_asset():
    wb = Workbook()
    ws = wb.active
    ws.title = "MultiAsset"
    ws.append(["Date", "Coal Consumption AFBC-1", "Coal Consumption AFBC-2", "Steam Generation AFBC-1", "Power Generation TG-1", "Power Generation TG-2", "Production Output VSF"])
    ws.append(["2026-02-20", 1200, 1180, 110.5, 85.2, 84.0, 500])
    ws.append(["2026-02-21", 1250, 1210, 112.0, 86.1, 84.7, 510])
    wb.save(OUT_DIR / "multi_asset.xlsx")


if __name__ == "__main__":
    make_clean_data()
    make_messy_data()
    make_multi_asset()
    print(f"Created: {OUT_DIR / 'clean_data.xlsx'}")
    print(f"Created: {OUT_DIR / 'messy_data.xlsx'}")
    print(f"Created: {OUT_DIR / 'multi_asset.xlsx'}")

